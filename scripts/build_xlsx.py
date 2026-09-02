#!/usr/bin/env python3
"""FRAZIYM Opportunity Database -> XLSX (9 sheets).
Follows xlsx skill: design.md tokens via templates/base.py, B2 origin,
borderless-first, Review cross-checks, computed values for programmatic QA.
"""
import sys, os, json
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)
from base import (FONT_NAME, HEADER_BOLD, PRIMARY, SECONDARY, NEUTRAL_900, NEUTRAL_600,
                  NEUTRAL_200, NEUTRAL_100, NEUTRAL_0, ACCENT_POSITIVE, ACCENT_NEGATIVE,
                  ACCENT_WARNING, font_title, font_header, font_subheader, font_body,
                  font_caption, fill_header, fill_total, fill_data_row, border_header,
                  border_total, align_title, align_header, align_number, align_text,
                  setup_sheet, style_header_row, style_data_row, style_total_row)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

OUT = "/home/z/my-project/download/FRAZIYM_Opportunity_Database.xlsx"
ROWS = json.load(open("/home/z/my-project/research/opportunities.json", encoding="utf-8"))

wb = Workbook()

# ---------------------------------------------------------------- helpers
def write_block(ws, start_row, headers, rows, widths=None, number_cols=None,
                align_num=None):
    """Write header row at start_row + data rows. Returns last row."""
    for c, h in enumerate(headers, 2):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, row_num=start_row, col_start=2, col_end=len(headers) + 1)
    r = start_row
    for i, row in enumerate(rows):
        r = start_row + 1 + i
        for c, v in enumerate(row, 2):
            cell = ws.cell(row=r, column=c, value=v)
            if align_num and (c - 2) in align_num:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        style_data_row(ws, row_num=r, col_start=2, col_end=len(headers) + 1, row_index=i)
        # re-apply number alignment after styling
        if align_num:
            for c in align_num:
                ws.cell(row=r, column=c + 2).alignment = Alignment(horizontal='right', vertical='center')
    if widths:
        for c, w in enumerate(widths, 2):
            ws.column_dimensions[get_column_letter(c)].width = w
    return r

def section_label(ws, row, text, last_col):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=PRIMARY)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 24

def caption(ws, row, text, last_col):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=2, value=text)
    c.font = font_caption()
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ---------------------------------------------------------------- 1 README
ws = wb.active
ws.title = "README"
setup_sheet(ws, title="FRAZIYM — One Product, Many Tools | Opportunity Database", last_col=8)
ws.sheet_view.showGridLines = False
r = 4
blocks = [
    ("What this file is", "Scored opportunity database for the FRAZIYM ecosystem: ONE subscription product that bundles 240+ micro-tools (including long-tail tools used <=1x/month), priced per user plan — never per-tool, never underpriced. Built from 7 research streams (2026-09-02) with URL-sourced evidence."),
    ("The 247 rows", "103 research-locked rows (from data-repair DR-*, e-commerce EC-*, devtools DV-*, weird-gold WG-* streams) + 144 long-tail catalog rows (GT-*, pattern-inherited E3, each pointing to its research parent). Every row = one tool/feature candidate for the platform."),
    ("Evidence legend", "E1 = direct community complaint/thread found (highest confidence). E2 = verified product/pricing/docs fact. E3 = pattern-inferred hypothesis — validate before building. 79 E1 / 15 E2 / 153 E3."),
    ("Scoring model (10 dims)", "Demand 20% | Pain 15% | Frequency 10% | Automation 15% | Monetization 10% | Distribution 10% | Competition 5% (high=low rivalry) | Infra-econ 5% | Ecosystem 5% | Defensibility 5%. Composite = weighted sum, 0-10."),
    ("Platform roles", "HOOK = free acquisition surface (SEO traffic). GLUE = keeps users returning (daily utilities). PRO = paywall-worthy professional features. LTV = flagship high-value tools that justify the subscription on their own (e.g., bank-statement PDF->CSV replaces $49.95/mo DocuClipper)."),
    ("Tier fit", "FREE = hook tools (client-side, unlimited) | PRO = full toolbox + batch + no watermark | MAX = teams/API/AI-heavy/server-metered features."),
    ("Anti-underpricing anchors", "Single-tool replacement costs verified in research: DocuClipper $49.95/mo, MoneyThumb $299.95, DataFeedWatch $59-64/mo, Matrixify $20-200/mo, Postman $14/user/mo, WP All Import $99-229/yr, Dext $13-20.50/mo. FRAZIYM PRO at $4.99 and STUDIO at $9.99 undercut every niche incumbent while covering all domains."),
    ("Sheets", "Opportunities = full 247-row filterable database. Top 50 = ranked shortlist. Clusters = zone-level stats. Pricing Model = subscription tiers. Revenue Scenarios = $0.3-$30/day paths. Payment Rails = BD-no-bank payout options. Free Infra + AI Providers = $0 stack limits. Review = live cross-checks."),
    ("Method note", "Scores are decision-support estimates combining evidence strength, parsed severity/frequency, automation class and verified pricing anchors — not survey data. Re-score before committing build slots. Sources per row live in research/opportunities.json (URLs included) and the Phase-1 research export."),
]
for label, text in blocks:
    section_label(ws, r, label, 8); r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    c = ws.cell(row=r, column=2, value=text)
    c.font = font_body(); c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[r].height = 46
    r += 2
caption(ws, r, "Prepared 2026-09-03 | Full sources: research/*.md + download/FRAZIYM_Research_Export_Phase1.md | Currency: USD", 8)
for col, w in zip("BCDEFGH", [16, 22, 22, 22, 22, 22, 22]):
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- 2 Opportunities
ws = wb.create_sheet("Opportunities")
setup_sheet(ws, title="All 247 Scored Opportunities (filterable)", last_col=25)
headers = ["Rank", "ID", "Tool / Opportunity", "Zone", "Cluster", "Evidence",
           "Composite", "Demand", "Pain", "Freq", "Automation", "Monetize",
           "Distrib", "Comp*", "Infra", "Eco", "Defen", "Role", "Tier",
           "Effort h", "Maint min/wk", "AI", "SEO query / parent", "Problem (abridged)"]
rows = []
for i, rec in enumerate(ROWS, 1):
    d = rec["dims"]
    problem = (rec.get("problem") or "")[:260]
    seo = rec.get("seo_query") or rec.get("notes") or ""
    rows.append([i, rec["id"], rec["title"], rec["zone"], rec["cluster"],
                 rec["ev_norm"], rec["composite"],
                 d["demand"], d["pain"], d["frequency"], d["automation"],
                 d["monetization"], d["distribution"], d["competition"],
                 d["infra_econ"], d["ecosystem"], d["defensibility"],
                 rec["platform_role"], rec["tier_fit"], rec["build_effort_h"],
                 rec["maintenance_min_wk"], "Y" if rec["ai_required"] else "N",
                 seo, problem])
last = write_block(ws, 4, headers, rows,
                   widths=[6, 15, 38, 14, 20, 9, 10, 8, 7, 7, 11, 9, 9, 8, 7, 6, 7, 7, 7, 8, 9, 5, 26, 60],
                   align_num=set(range(0, 17)) | {19, 20})
# composite color scale + dims color scale
comp_col = get_column_letter(8)  # column H = Composite
ws.conditional_formatting.add(f"{comp_col}5:{comp_col}{last}",
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B'))
for col in ["I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
    ws.conditional_formatting.add(f"{col}5:{col}{last}",
        ColorScaleRule(start_type='num', start_value=1, start_color='F8696B',
                       mid_type='num', mid_value=5.5, mid_color='FFEB84',
                       end_type='num', end_value=10, end_color='63BE7B'))
ws.auto_filter.ref = f"B4:Y{last}"
ws.freeze_panes = "E5"
caption(ws, last + 2, "*Comp = Competition score (10 = no/free incumbent, 1 = saturated category king). Composite = weighted 10-dim score. Full problem text + source URLs per row: research/opportunities.json. Evidence: E1 community-verified, E2 product/doc-verified, E3 pattern-inferred.", 25)

# ---------------------------------------------------------------- 3 Top 50
ws = wb.create_sheet("Top 50")
setup_sheet(ws, title="Top 50 Opportunities by Composite Score", last_col=11)
headers = ["Rank", "ID", "Tool / Opportunity", "Zone", "Score", "Role", "Tier",
           "Why it wins (abridged)", "Existing tools / price anchors", "Src URLs"]
rows = []
for i, rec in enumerate(ROWS[:50], 1):
    why = (rec.get("problem") or "")[:300]
    existing = (rec.get("existing_tools") or rec.get("evidence_summary") or "")[:220]
    rows.append([i, rec["id"], rec["title"], rec["zone"], rec["composite"],
                 rec["platform_role"], rec["tier_fit"], why, existing,
                 len(rec.get("sources") or [])])
last = write_block(ws, 4, headers, rows,
                   widths=[6, 15, 40, 14, 8, 7, 7, 60, 55, 9],
                   align_num={0, 4, 9})
ws.conditional_formatting.add(f"F5:F{last}",
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B'))
ws.auto_filter.ref = f"B4:K{last}"
ws.freeze_panes = "E5"

# ---------------------------------------------------------------- 4 Clusters
ws = wb.create_sheet("Clusters")
setup_sheet(ws, title="Zone-Level Portfolio Statistics", last_col=12)
zones = {}
for rec in ROWS:
    z = rec["zone"]
    zones.setdefault(z, []).append(rec)
headers = ["Zone", "Tools", "Avg score", "Max score", "LTV", "PRO", "HOOK", "GLUE",
           "E1 rows", "E2 rows", "Build effort h", "Maint min/wk"]
data = sorted(zones.items(), key=lambda kv: -sum(x["composite"] for x in kv[1]) / len(kv[1]))
rows = []
for z, recs in data:
    rows.append([z, len(recs), round(sum(x["composite"] for x in recs) / len(recs), 2),
                 max(x["composite"] for x in recs),
                 sum(1 for x in recs if x["platform_role"] == "LTV"),
                 sum(1 for x in recs if x["platform_role"] == "PRO"),
                 sum(1 for x in recs if x["platform_role"] == "HOOK"),
                 sum(1 for x in recs if x["platform_role"] == "GLUE"),
                 sum(1 for x in recs if x["ev_norm"] == "E1"),
                 sum(1 for x in recs if x["ev_norm"] == "E2"),
                 sum(x["build_effort_h"] for x in recs),
                 sum(x["maintenance_min_wk"] for x in recs)])
last = write_block(ws, 4, headers, rows,
                   widths=[24, 8, 10, 10, 7, 7, 8, 8, 9, 9, 13, 13],
                   align_num=set(range(1, 12)))
total_row = last + 1
ws.cell(row=total_row, column=2, value="TOTAL")
totals = [len(ROWS),
          round(sum(x["composite"] for x in ROWS) / len(ROWS), 2),
          max(x["composite"] for x in ROWS),
          sum(1 for x in ROWS if x["platform_role"] == "LTV"),
          sum(1 for x in ROWS if x["platform_role"] == "PRO"),
          sum(1 for x in ROWS if x["platform_role"] == "HOOK"),
          sum(1 for x in ROWS if x["platform_role"] == "GLUE"),
          sum(1 for x in ROWS if x["ev_norm"] == "E1"),
          sum(1 for x in ROWS if x["ev_norm"] == "E2"),
          sum(x["build_effort_h"] for x in ROWS),
          sum(x["maintenance_min_wk"] for x in ROWS)]
for c, v in enumerate(totals, 3):
    cell = ws.cell(row=total_row, column=c, value=v)
    cell.alignment = Alignment(horizontal='right', vertical='center')
style_total_row(ws, row_num=total_row, col_start=2, col_end=13)
caption(ws, total_row + 2, "Avg score column = mean composite per zone. Build effort = cumulative one-time build hours estimate (maint min/wk = recurring). Zones ordered by average score.", 13)

# ---------------------------------------------------------------- 5 Pricing Model
ws = wb.create_sheet("Pricing Model")
setup_sheet(ws, title="Subscription Model — One Product, Many Tools (anti-underpricing)", last_col=9)
headers = ["Tier", "Price", "Annual", "Who it serves", "Limits & unlockables",
           "Platform roles served", "Value anchor (what it replaces)", "Why price is fair"]
rows = [
    ["FREE", "$0", "$0", "Anyone arriving from search; students; casual one-off users",
     "All client-side tools unlimited; server/AI tools metered daily (10 credits); watermark/credit on shareable outputs; 1 concurrent job",
     "HOOK (53) + GLUE core (97)",
     "free alternatives: iLovePDF free, crontab.guru, jwt.io",
     "Free tier is the SEO engine: every tool page = a landing page; converts via caps + batch needs"],
    ["PRO", "$4.99/mo", "$39/yr", "Sellers, freelancers, students, office workers (the 95% case)",
     "Full 240+ toolbox; batch mode; no watermarks; 300 server credits/mo; 2GB files; priority queue",
     "PRO (65) + heavy GLUE",
     "iLovePDF ~$4-7/mo for PDF-only; Smallpdf $9/mo; Matrixify $20/mo for Shopify-only",
     "One niche incumbent charges more than PRO for ONE domain; FRAZIYM covers data+PDF+e-com+dev in one"],
    ["STUDIO", "$9.99/mo", "$79/yr", "Power users: accountants, agencies, AI-native devs",
     "PRO + AI-featured tools (1,500 AI credits/mo): citation verifier, error decoder, quiz/show-notes gen; API-lite 1k calls/mo; Telegram Mini App perks",
     "LTV (32) AI-assisted",
     "DocuClipper $49.95/mo (bank stmt only); Dext $13-20.50/mo (receipts only); DataFeedWatch $59/mo (feeds only)",
     "STUDIO replaces any ONE niche subscription at <25% of its price while including everything else"],
    ["MAX", "$19.99/mo", "$159/yr", "Small teams (3 seats), integrators, automation-heavy users",
     "STUDIO x3 seats + 5,000 AI credits + webhook workspace (bins/replay) + MCP tooling + 25k API calls + priority support",
     "LTV server-side + teams",
     "Postman $14/user/mo (1 user!); Hookdeck from ~$19/mo; ngrok $8-10/mo",
     "3-seat team plan beats Postman single-user price; metered server features protected by caps"],
]
last = write_block(ws, 4, headers, rows, widths=[10, 10, 9, 30, 46, 22, 40, 46])
r = last + 2
section_label(ws, r, "Profit mechanics", 9); r += 1
mech = [
    ["Infra cost", "$0", "Cloudflare free tier (Workers 100k req/day, Pages, D1, R2, Queues, Cron) — verified 2026-09-02"],
    ["AI cost", "$0", "Multi-provider free gateway: Groq + OpenRouter:free + CF Workers AI + SambaNova + Mistral ≈ 500-2,000 req/day pooled; cache-first routing"],
    ["Payment fees", "~2-5%", "USDT P2P->bKash ≈1.4-2% (cheapest); Telegram Stars ≈4-5% (21-day hold); Paddle 5%+$1 when scale requires"],
    ["Gross margin", "≈95-98%", "Revenue minus payment fees only; no hosting/AI fixed cost under free caps"],
    ["Underpricing guard", "Rule", "Any tier price change must stay above: (a) 2x cheapest single-niche competitor monthly price ÷ 20, (b) willingness-to-pay evidence floor $4/mo; never compete to $0"],
]
for row in mech:
    r += 1
    for c, v in enumerate(row, 2):
        ws.cell(row=r, column=c, value=v)
    style_data_row(ws, row_num=r, col_start=2, col_end=9, row_index=r - last - 3)
caption(ws, r + 2, "Prices in USD. Stars/crypto rails avoid the Stripe/PayPal wall for a Bangladesh solo founder (see Payment Rails sheet). Prices intentionally at market-mid of consumer utility SaaS, not race-to-bottom.", 9)

# ---------------------------------------------------------------- 6 Revenue Scenarios
ws = wb.create_sheet("Revenue Scenarios")
setup_sheet(ws, title="Revenue Scenarios — subscriber math (fees net of infra $0)", last_col=12)
# assumptions block
r = 4
section_label(ws, r, "Tier price assumptions (editable)", 12); r += 1
price_rows = [["PRO $/mo", 4.99], ["STUDIO $/mo", 9.99], ["MAX $/mo", 19.99],
              ["Ads+affiliate $/day (conservative)", 0.00]]
prow0 = r
for i, row in enumerate(price_rows):
    r += 1
    ws.cell(row=r, column=2, value=row[0])
    c = ws.cell(row=r, column=3, value=row[1]); c.number_format = '#,##0.00'
    style_data_row(ws, row_num=r, col_start=2, col_end=3, row_index=i)
pr = {"PRO": f"C{prow0+1}", "STUDIO": f"C{prow0+2}", "MAX": f"C{prow0+3}", "ADS": f"C{prow0+4}"}
r += 2
section_label(ws, r, "Scenarios (live formulas)", 12); r += 1
hdr_row = r
headers = ["Scenario", "PRO subs", "STUDIO subs", "MAX subs", "MRR $", "$/day",
           "Fees ~%", "Net $/day", "Funnel assumption (E3)", "Trigger date (90-day plan)", "Cumulative tools live"]
scen = [
    ["A — First blood", 1, 0, 0, 2, "Day 21-45: first paid conversion from one seeded tool's thank-you page", "Day 30: 6 tools", "6"],
    ["B — $1/day", 3, 1, 0, 3, "400-800 free-tool visits/day converting 0.5-1% to paid", "Day 45: 12 tools", "12"],
    ["C — $3/day", 8, 3, 1, 4, "1.5-3k visits/day via pSEO long-tail + Telegram channel", "Day 60: 25 tools", "25"],
    ["D — $10/day", 25, 8, 3, 4, "6-10k visits/day; churn <8%/mo; 2-4% free->paid on segment-fit tools", "Day 90: 45 tools", "45"],
    ["E — $30/day", 70, 25, 8, 5, "15-25k visits/day; email/Telegram list 5-10k; affiliate + AdsGram adds ~$7/day", "Month 5-6: 80+ tools", "80"],
]
rows = []
for s in scen:
    rows.append([s[0], s[1], s[2], s[3], None, None, s[4], None, s[5], s[6], s[7]])
last = write_block(ws, hdr_row, headers, rows,
                   widths=[18, 9, 11, 9, 10, 9, 8, 10, 44, 24, 12],
                   align_num={1, 2, 3, 4, 5, 6, 7})
# live formulas: MRR, $/day, net
for i in range(len(scen)):
    rr = hdr_row + 1 + i
    ws.cell(row=rr, column=6, value=f"=C{rr}*{pr['PRO']}+D{rr}*{pr['STUDIO']}+E{rr}*{pr['MAX']}").number_format = '#,##0.00'
    ws.cell(row=rr, column=7, value=f"=F{rr}*12/365").number_format = '#,##0.00'
    ws.cell(row=rr, column=9, value=f"=G{rr}*(1-H{rr}/100)+{pr['ADS']}").number_format = '#,##0.00'
# fee percents column H values
for i, pct in enumerate([2.0, 2.5, 3.0, 4.0, 5.0]):
    rr = hdr_row + 1 + i
    c = ws.cell(row=rr, column=8, value=pct); c.number_format = '0.0'
    c.alignment = Alignment(horizontal='right', vertical='center')
# fix scenario C,D,E ads assumptions: put ads into ADS cell? keep 0 base; note in caption
caption(ws, last + 2, "MRR and $/day are live formulas — edit tier prices above to re-model. Fees = blended payment-rail cost (see Payment Rails). Scenario D/E assume affiliate + AdsGram top-ups beyond subscriptions (entered conservatively at $0 here; PDF report shows with +$1.5-7/day). Free->paid conversion assumptions are E3 and must be validated with real funnel data. Note: '工具-fit' = segment fit.", 12)
ws.freeze_panes = "C6"

# ---------------------------------------------------------------- 7 Payment Rails
ws = wb.create_sheet("Payment Rails")
setup_sheet(ws, title="Payout Rails for BD Solo Founder (no bank/card) — verified 2026-09-02", last_col=10)
headers = ["Rail", "Verdict", "Fees", "Min payout", "KYC", "Timeline", "Notes / risk", "Confidence"]
rows = [
    ["Crypto affiliate USDT -> P2P -> bKash (Bybit/OKX)", "WORKS — cheapest chain", "~1.4-2% total", "exchange min", "Exchange KYC + P2P counterparty", "hours-1 day", "Prefer Bybit/OKX P2P (0% fee, bKash/Nagad listed); Binance BD restricted Mar 2025; USDT/BDT rate 122-124 = +2-3.5% FX bonus; risk = MFS wallet freezes from many small P2P credits", "High (E2)"],
    ["Telegram Stars -> Fragment -> GRAM -> USDT -> P2P", "WORKS — native for TG products", "~4-5% total", "1,000 Stars (~$13)", "Fragment KYC mandatory (purchases since Nov 2024)", "21-day hold + payout", "$0.013/Star payout rate; Toncoin renamed GRAM 1:1 (Jun 15, 2026)", "High (E2)"],
    ["AdsGram (TGA mini-app ads)", "WORKS — official TG ad network", "net-45 style", "$100", "light", "monthly", "USDT-TON payouts; rewarded ads in Mini Apps", "Med-High (E2)"],
    ["Adsterra", "WORKS (crypto)", "rev-share", "$100 crypto (USDT-TRC20)", "email", "net-14/16", "The famous $5 min is WebMoney/Paxum only, NOT crypto", "High (E2)"],
    ["Paddle -> Payoneer -> bKash", "WORKS — only mainstream SaaS checkout open to BD", "3% + $1 (Payoneer->bKash min BDT 1000, cap BDT 250k/tx)", "Paddle $1,000 initial review bar", "Payoneer KYC (BD ID ok)", "weekly/2x-monthly", "Merchant-of-record solves sales tax; use when subscription volume justifies", "Med-High (E2)"],
    ["NOWPayments / BTCPay (accept rails)", "WORKS — inbound crypto acceptance", "0.5-1.5% / 0% (self-host BTC-only)", "none-material", "email", "instant", "Accept USDT/BTC from customers; pair with P2P off-ramp", "Med (E2)"],
    ["Direct MTO remittance -> bKash (WU/MoneyGram/ACE/Taptap)", "WORKS but compliance-grey for client fees", "~0.7% ATM cash-out; +2.5% govt incentive on remittance", "low", "ID per transfer", "minutes-hours", "Cheapest per-dollar; grey area for commercial receipts (FERA) — use for personal withdrawals, not client invoicing", "Med (E2/E3)"],
    ["Google AdSense", "BROKEN", "-", "$100", "-", "-", "BD = wire-to-bank only; WU Quick Cash dead", "High (E2)"],
    ["Stripe / PayPal / Gumroad / Ko-fi / BMC / Patreon / Payhip / LemonSqueezy", "BROKEN", "-", "-", "-", "-", "All require BD bank or Stripe/PayPal payout country", "High (E2)"],
    ["Telegram channel ads rev-share (>=1k subs, 50%)", "UNVERIFIED for BD", "50% share", "-", "-", "-", "Check in-app toggle; AdsGram is the fallback", "Low (E3)"],
    ["Bangladesh Bank legal note", "CONSTRAINT", "-", "-", "-", "-", "Crypto = grey zone (FERA 1947); practical risk = MFS freezes, not prosecution; keep P2P credits small & irregular", "Med (E2/E3)"],
]
last = write_block(ws, 4, headers, rows, widths=[38, 22, 20, 16, 22, 14, 60, 12])
caption(ws, last + 2, "Full ledger with sources: research/payments.md. Strategy: start Stars + crypto (zero-setup), graduate to Paddle->Payoneer at scale for card-paying customers.", 10)

# ---------------------------------------------------------------- 8 Free Infra
ws = wb.create_sheet("Free Infra")
setup_sheet(ws, title="$0 Infrastructure Stack — verified limits 2026-09-02", last_col=7)
headers = ["Service", "Free-tier allowance", "Hard gotcha", "Role in FRAZIYM"]
rows = [
    ["Cloudflare Workers", "100k req/day, 10ms CPU", "CPU-time cap kills heavy compute; offload to client WASM", "API gateway + router + bots"],
    ["Cloudflare Pages", "unlimited bandwidth, 500 builds/mo", "—", "all static tool frontends"],
    ["Cloudflare D1", "5M reads/day, 100k writes/day, 5GB", "writes capped — batch/analytics offload", "users, usage ledger, tool registry"],
    ["Cloudflare R2", "10GB storage, 1M A/10M B ops, $0 egress", "—", "file staging for server-side jobs"],
    ["Cloudflare Queues + Cron", "Queues 10k ops/day (Feb 2026 change); Cron reliable", "—", "async jobs + schedulers"],
    ["Cloudflare KV", "1k writes/day", "low-write only", "config/cache"],
    ["Turso / Neon", "Turso 500M reads/10M writes/5GB; Neon 100 CU-hr + 5-min autosuspend", "Neon autosuspend = cold starts; heartbeat cron mandatory", "DB depth beyond D1"],
    ["Supabase", "2 projects, 500MB", "PAUSES after 7 days idle", "avoid as primary"],
    ["GitHub", "Actions 2,000 min/mo private (unlimited public); scheduled workflows BEST-EFFORT (60-day disable)", "never rely on GH cron for revenue jobs", "CI, OSS repo trust, Pages mirror"],
    ["Vercel Hobby", "100GB, 1M invocations", "NON-COMMERCIAL terms; cron max 1/day", "avoid for the product itself"],
    ["Netlify", "300 credits (~15GB) hard cap", "downgraded free tier", "dropped from stack"],
    ["Resend + Brevo (email)", "Resend 100/day + 3,000/mo; Brevo 300/day", "custom domain required for deliverability", "transactional email dual-rail"],
    ["cron-job.org / QStash / healthchecks.io", "unlimited jobs @1-min / 1,000 msgs/day / 20 monitors", "—", "external schedulers + dead-man watchdog"],
    ["UptimeRobot / Better Stack", "50 monitors x 5min (non-commercial wording) / 10", "—", "monitoring"],
    ["Clerk / Auth0 / Supabase Auth", "10k MAU / 25k MAU / 50k MAU", "Clerk 10k claim unverified", "auth at launch = Telegram + magic link"],
    ["Cloudflare Web Analytics / PostHog", "free / 1M events/mo", "—", "funnel analytics"],
    ["Domains", "DigitalPlat free (.dpdns.org/.us.kg); .xyz/.top $1-2 yr-1", "custom domain REQUIRED for email/OAuth trust", "root domain + API subdomain"],
]
last = write_block(ws, 4, headers, rows, widths=[30, 40, 40, 34])
caption(ws, last + 2, "Full 16-section verification with sources: research/infra.md. Architecture rule: anchor on Cloudflare; treat every free tier as disposable (failover-ready).", 7)

# ---------------------------------------------------------------- 9 AI Providers
ws = wb.create_sheet("AI Providers")
setup_sheet(ws, title="Free AI Inference Providers — gateway design inputs 2026-09-02", last_col=7)
headers = ["Provider", "Free allowance (verified)", "Card needed", "ToS / data notes", "Role in router"]
rows = [
    ["Groq", "30 RPM / 1K RPD / 8K TPM (gpt-oss + qwen); Whisper 2K RPD & 28.8k audio-sec/day", "No", "cached tokens excluded from limits", "PRIMARY anchor (speed + STT)"],
    ["OpenRouter :free", "20 RPM, 50 RPD (1,000 RPD after $10 credit purchase)", "No", "free models = data-training pool", "SECONDARY + model variety"],
    ["Cloudflare Workers AI", "10,000 Neurons/day", "No", "same stack as product = one bill, zero", "TERTIARY + embedding/small tasks"],
    ["SambaNova", "20 RPM / 20 RPD / 200K TPD", "No", "—", "failover"],
    ["Mistral La Plateforme", "free tier exists (rate-limited)", "Phone verify", "—", "failover"],
    ["Cohere", "trial 1,000 calls/mo @20 RPM", "No", "trial key only", "embed/rank niche"],
    ["Gemini (AI Studio)", "free tier exists but ~20 RPD (Dec 2025 slash; no public table)", "No", "DATA USED FOR TRAINING = YES", "last-resort only"],
    ["Hugging Face Inference", "~$0.10/mo free credits", "No", "—", "drop from router"],
    ["Cerebras / Together / GitHub Models", "ALL free tiers RETIRED (2025-07/2026-02/2026-07)", "-", "—", "removed"],
    ["Local (llama.cpp on Termux)", "0-cost, device-bound", "-", "—", "offline dev-time fallback only"],
]
last = write_block(ws, 4, headers, rows, widths=[24, 44, 12, 34, 28])
caption(ws, last + 2, "Router policy: validate -> deterministic first -> cache -> small/local model -> free-cloud failover -> validate again. Pooled capacity ≈ 500-2,000 AI req/day at $0. Single account per provider (multi-accounting prohibited). Sources: research/ai-providers.md.", 7)

# ---------------------------------------------------------------- 10 Review
ws = wb.create_sheet("Review")
ws.sheet_properties.tabColor = "FFC000"
setup_sheet(ws, title="Cross-Validation (live checks)", last_col=6)
opp_last = 4 + len(ROWS)  # Opportunities data ends here
checks = [
    ["Check", "Expected", "Actual", "Status"],
    ["Opportunities row count", len(ROWS), f"=COUNTA(Opportunities!C5:C{opp_last})", None],
    ["Top 50 row count", 50, "=COUNTA('Top 50'!C5:C54)", None],
    ["Clusters tool sum", len(ROWS), "=Clusters!C12", None],
    ["Avg composite (rounded)", round(sum(x["composite"] for x in ROWS) / len(ROWS), 2),
     f"=ROUND(AVERAGE(Opportunities!H5:H{opp_last}),2)", None],
    ["Scenario B MRR $", "=3*Pricing_Model_Range+0", None, None],  # placeholder replaced below
]
# simpler: build checks without the placeholder row
checks = checks[:4] + [
    ["Scenario B MRR ($)", 24.96, "='Revenue Scenarios'!F14", None],
    ["LTV+PRO+HOOK+GLUE sum", len(ROWS),
     f"=COUNTIF(Opportunities!S5:S{opp_last},\"LTV\")+COUNTIF(Opportunities!S5:S{opp_last},\"PRO\")+COUNTIF(Opportunities!S5:S{opp_last},\"HOOK\")+COUNTIF(Opportunities!S5:S{opp_last},\"GLUE\")", None],
]
r = 4
for c, h in enumerate(checks[0], 2):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, row_num=r, col_start=2, col_end=5)
for i, row in enumerate(checks[1:]):
    r += 1
    ws.cell(row=r, column=2, value=row[0])
    ws.cell(row=r, column=3, value=row[1])
    ws.cell(row=r, column=4, value=row[2])
    ws.cell(row=r, column=5, value=f'=IF(ROUND(C{r},2)=ROUND(D{r},2),"PASS","FAIL")')
    style_data_row(ws, row_num=r, col_start=2, col_end=5, row_index=i)
for col, w in zip("BCDE", [30, 16, 18, 12]):
    ws.column_dimensions[col].width = w

wb.properties.creator = "Z.ai"
wb.properties.title = "FRAZIYM Opportunity Database"
wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)
